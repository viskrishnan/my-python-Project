import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
Product_list = inventory_file["Sheet1"]

products_per_supplier = {}
for product_row in range(2, Product_list.max_row + 1):
    supplier_name = Product_list.cell(product_row, 4).value
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

print(products_per_supplier)